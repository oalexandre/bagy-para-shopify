#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para comparar produtos da Shopify com produtos da Bagy
e gerar um arquivo Excel com as correspondências encontradas.

Busca produtos da API Shopify e compara com produtos.json da Bagy
gerando um relatório com ID Shopify, ID Bagy, URL Shopify e URL Bagy.
"""

import json
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import time
from difflib import SequenceMatcher
from dotenv import load_dotenv
import re

# Carrega variáveis de ambiente
load_dotenv()

def clean_string_for_comparison(text):
    """Remove caracteres especiais e normaliza texto para comparação"""
    if not text:
        return ""
    
    # Remove acentos
    text = text.lower()
    text = text.replace('ã', 'a').replace('á', 'a').replace('à', 'a').replace('â', 'a')
    text = text.replace('é', 'e').replace('ê', 'e').replace('í', 'i').replace('ó', 'o')
    text = text.replace('ô', 'o').replace('õ', 'o').replace('ú', 'u').replace('ç', 'c')
    
    # Remove caracteres especiais, mantém apenas letras, números e espaços
    text = re.sub(r'[^a-z0-9\s]', '', text)
    
    # Remove espaços extras
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text

def calculate_similarity(text1, text2):
    """Calcula a similaridade entre dois textos"""
    clean1 = clean_string_for_comparison(text1)
    clean2 = clean_string_for_comparison(text2)
    
    return SequenceMatcher(None, clean1, clean2).ratio()

def get_shopify_products(shop_domain, access_token):
    """Busca todos os produtos da loja Shopify"""
    products = []
    page_info = None
    
    print("Buscando produtos da Shopify...")
    
    while True:
        url = f"https://{shop_domain}/admin/api/2023-10/products.json"
        
        params = {
            'limit': 250,  # Máximo permitido pela API
            'fields': 'id,title,handle,created_at,updated_at,status'
        }
        
        if page_info:
            params['page_info'] = page_info
            
        headers = {
            'X-Shopify-Access-Token': access_token,
            'Content-Type': 'application/json'
        }
        
        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            
            data = response.json()
            batch_products = data.get('products', [])
            
            if not batch_products:
                break
                
            products.extend(batch_products)
            print(f"Carregados {len(products)} produtos da Shopify...")
            
            # Verifica se há mais páginas
            link_header = response.headers.get('Link')
            if link_header and 'rel="next"' in link_header:
                # Extrai page_info do cabeçalho Link
                next_link = [link for link in link_header.split(',') if 'rel="next"' in link][0]
                page_info = next_link.split('page_info=')[1].split('&')[0].split('>')[0]
            else:
                break
                
        except requests.exceptions.RequestException as e:
            print(f"Erro ao buscar produtos da Shopify: {e}")
            break
            
        # Delay para evitar rate limiting
        time.sleep(0.5)
    
    print(f"Total de produtos encontrados na Shopify: {len(products)}")
    return products

def load_bagy_products(json_file_path):
    """Carrega produtos do arquivo JSON da Bagy"""
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            products = json.load(f)
        
        # Filtra produtos válidos
        valid_products = [p for p in products if p and p.get('name') and p.get('id')]
        
        print(f"Produtos válidos carregados da Bagy: {len(valid_products)}")
        return valid_products
    
    except Exception as e:
        print(f"Erro ao carregar produtos da Bagy: {e}")
        return []

def find_matching_products(shopify_products, bagy_products, similarity_threshold=0.7):
    """Encontra produtos correspondentes entre Shopify e Bagy"""
    matches = []
    
    print(f"Comparando produtos (similaridade mínima: {similarity_threshold})...")
    
    for i, shopify_product in enumerate(shopify_products):
        shopify_title = shopify_product.get('title', '')
        best_match = None
        best_similarity = 0
        
        for bagy_product in bagy_products:
            bagy_name = bagy_product.get('name', '')
            similarity = calculate_similarity(shopify_title, bagy_name)
            
            if similarity > best_similarity and similarity >= similarity_threshold:
                best_similarity = similarity
                best_match = bagy_product
        
        if best_match:
            shop_domain = os.getenv('SHOPIFY_SHOP_DOMAIN', 'sua-loja')
            shopify_url = f"https://{shop_domain}/products/{shopify_product.get('handle', '')}"
            bagy_url = best_match.get('url', '')
            
            match = {
                'shopify_id': shopify_product.get('id'),
                'bagy_id': best_match.get('id'),
                'shopify_title': shopify_title,
                'bagy_name': best_match.get('name'),
                'shopify_url': shopify_url,
                'bagy_url': bagy_url,
                'similarity': best_similarity
            }
            
            matches.append(match)
            print(f"Match encontrado: {shopify_title} <-> {best_match.get('name')} (similaridade: {best_similarity:.2f})")
        
        # Progresso
        if (i + 1) % 50 == 0:
            print(f"Processados {i + 1}/{len(shopify_products)} produtos...")
    
    print(f"Total de correspondências encontradas: {len(matches)}")
    return matches

def create_excel_report(matches, output_path):
    """Cria arquivo Excel com as correspondências encontradas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Correspondências Shopify x Bagy"
    
    # Cabeçalhos
    headers = [
        'ID Shopify',
        'ID Bagy', 
        'Título Shopify',
        'Nome Bagy',
        'URL Shopify',
        'URL Bagy',
        'Similaridade'
    ]
    
    # Estilo para cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Adiciona cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Adiciona dados
    for row, match in enumerate(matches, 2):
        ws.cell(row=row, column=1, value=match['shopify_id'])
        ws.cell(row=row, column=2, value=match['bagy_id'])
        ws.cell(row=row, column=3, value=match['shopify_title'])
        ws.cell(row=row, column=4, value=match['bagy_name'])
        ws.cell(row=row, column=5, value=match['shopify_url'])
        ws.cell(row=row, column=6, value=match['bagy_url'])
        ws.cell(row=row, column=7, value=f"{match['similarity']:.2f}")
    
    # Ajusta largura das colunas
    column_widths = [15, 15, 40, 40, 60, 60, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Salva arquivo
    wb.save(output_path)
    print(f"Relatório Excel salvo em: {output_path}")

def main():
    """Função principal"""
    # Configurações da API Shopify
    shop_domain = os.getenv('SHOPIFY_SHOP_DOMAIN')
    access_token = os.getenv('SHOPIFY_ACCESS_TOKEN')
    
    if not shop_domain or not access_token:
        print("❌ Erro: Configurações da Shopify não encontradas!")
        print("Configure as variáveis SHOPIFY_SHOP_DOMAIN e SHOPIFY_ACCESS_TOKEN no arquivo .env")
        print("\nExemplo do arquivo .env:")
        print("SHOPIFY_SHOP_DOMAIN=sua-loja.myshopify.com")
        print("SHOPIFY_ACCESS_TOKEN=seu_token_de_acesso")
        return
    
    # Caminho do arquivo JSON da Bagy
    bagy_json_path = os.path.join("imported", "produtos.json")
    
    if not os.path.exists(bagy_json_path):
        print(f"❌ Arquivo não encontrado: {bagy_json_path}")
        print("Execute primeiro o script importProductsFromBagy.py para gerar o arquivo produtos.json")
        return
    
    # Carrega produtos
    print("=== COMPARAÇÃO DE PRODUTOS SHOPIFY x BAGY ===")
    shopify_products = get_shopify_products(shop_domain, access_token)
    bagy_products = load_bagy_products(bagy_json_path)
    
    if not shopify_products or not bagy_products:
        print("❌ Erro: Não foi possível carregar os produtos")
        return
    
    # Encontra correspondências
    matches = find_matching_products(shopify_products, bagy_products)
    
    if not matches:
        print("❌ Nenhuma correspondência encontrada")
        return
    
    # Cria pasta converted se não existir
    output_dir = "converted"
    os.makedirs(output_dir, exist_ok=True)
    
    # Gera arquivo Excel
    output_path = os.path.join(output_dir, "correspondencias_shopify_bagy.xlsx")
    create_excel_report(matches, output_path)
    
    print(f"\n✅ Comparação concluída!")
    print(f"📊 Produtos Shopify: {len(shopify_products)}")
    print(f"📊 Produtos Bagy: {len(bagy_products)}")
    print(f"🔗 Correspondências encontradas: {len(matches)}")
    print(f"📄 Relatório salvo em: {output_path}")

if __name__ == "__main__":
    main()
