import requests
import openpyxl
import os
from dotenv import load_dotenv
from datetime import datetime
import json
from time import sleep

load_dotenv()

SHOPIFY_SHOP_DOMAIN = os.getenv("SHOPIFY_SHOP_DOMAIN")
SHOPIFY_ACCESS_TOKEN = os.getenv("SHOPIFY_ACCESS_TOKEN")
SHOPIFY_API_VERSION = "2024-10"

def read_excel_coupons(filename="imported/cupons_dooca.xlsx"):
    if not os.path.exists(filename):
        print(f"❌ Arquivo {filename} não encontrado!")
        return []
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    coupons = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        coupon_data = dict(zip(headers, row))
        if coupon_data.get('active'):
            # Se não há código, gerar um baseado no nome
            if not coupon_data.get('codes'):
                name = coupon_data.get('name', '')
                # Criar código a partir do nome: remover caracteres especiais e usar uppercase
                code = ''.join(c for c in name if c.isalnum() or c.isspace())
                code = code.replace(' ', '').upper()[:20]  # Limitar a 20 caracteres
                if not code:
                    code = f"CUPOM{coupon_data.get('id', '')}"
                coupon_data['codes'] = code
            coupons.append(coupon_data)
    
    print(f"📊 {len(coupons)} cupons ativos encontrados no arquivo Excel")
    return coupons

def convert_bagy_to_shopify_format(bagy_coupon):
    shopify_discount = {
        "price_rule": {
            "title": bagy_coupon.get('name', ''),
            "target_type": "line_item",
            "target_selection": "all",
            "allocation_method": "across",
            "value_type": "percentage" if bagy_coupon.get('value_type') == 'percentage' else "fixed_amount",
            "value": f"-{float(bagy_coupon.get('value', 0))}",
            "customer_selection": "all",
            "once_per_customer": bagy_coupon.get('single_usage', False),
            "usage_limit": bagy_coupon.get('usage_limit') if bagy_coupon.get('usage_limit') else None,
            "starts_at": bagy_coupon.get('date_from'),
            "ends_at": bagy_coupon.get('date_to')
        }
    }
    
    if bagy_coupon.get('min_purchase'):
        shopify_discount["price_rule"]["prerequisite_subtotal_range"] = {
            "greater_than_or_equal_to": str(bagy_coupon.get('min_purchase'))
        }
    
    if bagy_coupon.get('min_quantity'):
        shopify_discount["price_rule"]["prerequisite_quantity_range"] = {
            "greater_than_or_equal_to": bagy_coupon.get('min_quantity')
        }
    
    if bagy_coupon.get('prerequisite_product_ids'):
        product_ids = [pid.strip() for pid in str(bagy_coupon.get('prerequisite_product_ids')).split(',') if pid.strip()]
        if product_ids:
            shopify_discount["price_rule"]["prerequisite_product_ids"] = product_ids
    
    if bagy_coupon.get('entitled_product_ids'):
        product_ids = [pid.strip() for pid in str(bagy_coupon.get('entitled_product_ids')).split(',') if pid.strip()]
        if product_ids:
            shopify_discount["price_rule"]["entitled_product_ids"] = product_ids
    
    return shopify_discount

def create_price_rule(discount_data):
    url = f"https://{SHOPIFY_SHOP_DOMAIN}/admin/api/{SHOPIFY_API_VERSION}/price_rules.json"
    headers = {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json"
    }
    
    response = requests.post(url, headers=headers, json=discount_data)
    
    if response.status_code == 201:
        return response.json()["price_rule"]
    else:
        print(f"❌ Erro ao criar price rule: {response.status_code}")
        print(f"   Resposta: {response.text}")
        return None

def create_discount_code(price_rule_id, code):
    url = f"https://{SHOPIFY_SHOP_DOMAIN}/admin/api/{SHOPIFY_API_VERSION}/price_rules/{price_rule_id}/discount_codes.json"
    headers = {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json"
    }
    
    discount_code_data = {
        "discount_code": {
            "code": code
        }
    }
    
    response = requests.post(url, headers=headers, json=discount_code_data)
    
    if response.status_code == 201:
        return response.json()["discount_code"]
    else:
        print(f"❌ Erro ao criar código de desconto: {response.status_code}")
        print(f"   Resposta: {response.text}")
        return None

def import_coupons_to_shopify():
    coupons = read_excel_coupons()
    
    if not coupons:
        print("❌ Nenhum cupom para importar")
        return
    
    success_count = 0
    error_count = 0
    results = []
    
    print(f"\n🚀 Iniciando importação de {len(coupons)} cupons para Shopify...")
    
    for i, coupon in enumerate(coupons, 1):
        print(f"\n[{i}/{len(coupons)}] Processando cupom: {coupon.get('name')} - Código: {coupon.get('codes')}")
        
        try:
            shopify_discount = convert_bagy_to_shopify_format(coupon)
            
            price_rule = create_price_rule(shopify_discount)
            
            if price_rule:
                discount_code = create_discount_code(price_rule["id"], coupon.get('codes'))
                
                if discount_code:
                    print(f"   ✅ Cupom '{coupon.get('name')}' importado com sucesso!")
                    success_count += 1
                    results.append({
                        "bagy_id": coupon.get('id'),
                        "bagy_code": coupon.get('codes'),
                        "shopify_price_rule_id": price_rule["id"],
                        "shopify_discount_code_id": discount_code["id"],
                        "status": "success"
                    })
                else:
                    print(f"   ⚠️ Price rule criada mas falha ao criar código de desconto")
                    error_count += 1
                    results.append({
                        "bagy_id": coupon.get('id'),
                        "bagy_code": coupon.get('codes'),
                        "status": "error",
                        "error": "Falha ao criar código de desconto"
                    })
            else:
                print(f"   ❌ Falha ao criar price rule")
                error_count += 1
                results.append({
                    "bagy_id": coupon.get('id'),
                    "bagy_code": coupon.get('codes'),
                    "status": "error",
                    "error": "Falha ao criar price rule"
                })
                
        except Exception as e:
            print(f"   ❌ Erro ao processar cupom: {str(e)}")
            error_count += 1
            results.append({
                "bagy_id": coupon.get('id'),
                "bagy_code": coupon.get('codes'),
                "status": "error",
                "error": str(e)
            })
        
        sleep(0.5)
    
    save_import_results(results)
    
    print(f"\n{'='*50}")
    print(f"📊 RESUMO DA IMPORTAÇÃO:")
    print(f"   ✅ Cupons importados com sucesso: {success_count}")
    print(f"   ❌ Cupons com erro: {error_count}")
    print(f"   📁 Relatório salvo em: imported/import_results.json")
    print(f"{'='*50}")

def save_import_results(results):
    os.makedirs("imported", exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"imported/import_results_{timestamp}.json"
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    with open("imported/import_results.json", 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n📝 Resultados salvos em {filename}")

def check_existing_discount_codes():
    url = f"https://{SHOPIFY_SHOP_DOMAIN}/admin/api/{SHOPIFY_API_VERSION}/price_rules.json"
    headers = {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json"
    }
    
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        price_rules = response.json()["price_rules"]
        print(f"📊 {len(price_rules)} price rules existentes na Shopify")
        return price_rules
    else:
        print(f"❌ Erro ao verificar price rules existentes: {response.status_code}")
        return []

def test_import_single_coupon():
    """Importa apenas o primeiro cupom para teste"""
    coupons = read_excel_coupons()
    
    if not coupons:
        print("❌ Nenhum cupom para importar")
        return
    
    # Pegar apenas o primeiro cupom para teste
    test_coupon = coupons[0]
    print(f"\n🧪 TESTE: Importando cupom: {test_coupon.get('name')}")
    print(f"   Código: {test_coupon.get('codes')}")
    
    try:
        shopify_discount = convert_bagy_to_shopify_format(test_coupon)
        print(f"   Formato Shopify: {json.dumps(shopify_discount, indent=2)}")
        
        price_rule = create_price_rule(shopify_discount)
        
        if price_rule:
            discount_code = create_discount_code(price_rule["id"], test_coupon.get('codes'))
            
            if discount_code:
                print(f"   ✅ Cupom de teste importado com sucesso!")
                print(f"   Price Rule ID: {price_rule['id']}")
                print(f"   Discount Code ID: {discount_code['id']}")
            else:
                print(f"   ❌ Falha ao criar código de desconto")
        else:
            print(f"   ❌ Falha ao criar price rule")
            
    except Exception as e:
        print(f"   ❌ Erro: {str(e)}")

if __name__ == "__main__":
    print("🛍️ IMPORTADOR DE CUPONS BAGY → SHOPIFY")
    print("="*50)
    
    print("\n1. Importar cupons do Excel para Shopify")
    print("2. Verificar cupons existentes na Shopify")
    print("3. Testar importação de 1 cupom")
    print("4. Sair")
    
    choice = input("\nEscolha uma opção: ")
    
    if choice == "1":
        import_coupons_to_shopify()
    elif choice == "2":
        existing = check_existing_discount_codes()
        for rule in existing[:5]:
            print(f"  - {rule.get('title')} (ID: {rule.get('id')})")
        if len(existing) > 5:
            print(f"  ... e mais {len(existing) - 5} cupons")
    elif choice == "3":
        test_import_single_coupon()
    else:
        print("👋 Encerrando...")